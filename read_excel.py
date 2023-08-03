from openpyxl import load_workbook
import os
from openpyxl.chart import BarChart, Reference

excel_folder = os.path.join(os.getcwd(), 'excel_file')
excel_name = 'data_no_graph.xlsx'
excel_path = os.path.join(excel_folder, excel_name)

excel_file = load_workbook(excel_path)
sheet = excel_file.active


print('row: ', sheet.max_row)
print('column: ', sheet.max_column)

value = Reference(sheet, min_col=3, max_col=3, min_row=2, max_row=4)
cat = Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=4)

chart = BarChart()
chart.title = "ความสูงของต้นไม้"
chart.y_axis.title = 'ความสูงของต้นไม้ (เมตร)'
chart.x_axis.title = 'ชนิดของต้นไม้'
chart.legend = None




chart.add_data(value)
sheet.add_chart(chart, 'E5')

chart.set_categories(cat)

excel_file.save(excel_path)